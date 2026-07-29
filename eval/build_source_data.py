"""Assemble ALL figure/result raw data into one NMI-style Source Data workbook:
data/source_data.xlsx  — one sheet per figure/result, editable, submission-ready.
Sources: curated CSVs (diagnosis/ladder/mitigation/R5) + drift_<m>.json + condsent_<m>.json.
Run: MOLREHALLU_REGEN=1 python eval/build_source_data.py
"""
import glob, json, os
import pandas as pd
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
D = os.path.join(BASE, "data")
AO = os.path.join(BASE, "data", "raw")
XLSX = os.path.join(D, "source_data.xlsx")

# Guard: this script REBUILDS data/source_data.xlsx in place, and it needs inputs the public
# release does not ship (data/token_examples/*.csv, the full data/raw/ set). Run against the
# released subset it would overwrite the submitted workbook with fewer, incomplete sheets.
# It also writes at import time, so `import build_source_data` alone is destructive.
if os.environ.get("MOLREHALLU_REGEN") != "1":
    raise SystemExit(
        "build_source_data.py rebuilds data/source_data.xlsx from the full evaluation tree, "
        "which is not part of the public release. Refusing to run so the shipped workbook is "
        "not overwritten with incomplete data. Set MOLREHALLU_REGEN=1 to override."
    )

# file-stem -> paper label (order = ladder order); exclude DeepSeek (unusable)
LABEL = {"Llama-3.1-8B-Instruct-base": "base-a (pre-SFT)", "Chem-R-SFT": "SFT",
         "+process": "+process", "Chem-R-Faithful": "+coupled", "Chem-R": "Chem-R",
         "ChemDFM-R": "ChemDFM-R", "ether-0": "ether-0"}
CONDS = ["syn_cot", "wrong_cot", "all_wrong_cot", "drop_cot", "swap_cot", "wrong_input"]
COND_DESC = {"syn_cot": "synonym (control)", "wrong_cot": "1 FG-name corrupted",
             "all_wrong_cot": "all FG-name claims corrupted", "drop_cot": "empty CoT (presence, OOD)",
             "swap_cot": "another molecule's whole CoT", "wrong_input": "FG corrupted in INPUT",
             "mask_draft": "drafted SMILES -> [...] (draft presence)",
             "corrupt_draft": "drafted SMILES -> valid-but-wrong structure (draft content)"}


def models_present(prefix):
    out = []
    for f in sorted(glob.glob(f"{AO}/{prefix}_*.json")):
        stem = os.path.basename(f)[len(prefix) + 1:-5]
        if stem in LABEL:
            out.append((stem, f))
    out.sort(key=lambda x: list(LABEL).index(x[0]))
    return out


PE_CONDS = ["syn_cot", "wrong_cot", "all_wrong_cot", "drop_cot", "swap_cot"]  # in per_example
TRANSLATE = {"cap2mol", "mol2cap", "retrosynthesis"}


def _flip(sub, c):
    # PAPER metric: among originally-correct, fraction that became wrong (dperf==-1).
    # NaN if the condition was not run for this model (key absent) — not 0%.
    present = [e for e in sub if f"{c}_dperf" in e]
    if not present:
        return float("nan")
    return sum(1 for e in present if e[f"{c}_dperf"] == -1) / len(present)


def drift_frames():
    grp = lambda t: "translate" if t in TRANSLATE else "s2"
    over, byt = [], []
    for stem, f in models_present("drift"):
        d = json.load(open(f)); lab = LABEL[stem]; pe = d["per_example"]
        for g in ["translate", "s2"]:
            sub = [e for e in pe if grp(e["task"]) == g and e.get("base_correct") == 1]
            for c in PE_CONDS:
                over.append({"model": lab, "task_group": g, "condition": c, "meaning": COND_DESC[c],
                             "flip_to_wrong_pct": round(_flip(sub, c) * 100, 2), "n_orig_correct": len(sub)})
        for t in sorted({e["task"] for e in pe}):
            sub = [e for e in pe if e["task"] == t and e.get("base_correct") == 1]
            for c in PE_CONDS:
                byt.append({"model": lab, "task": t, "condition": c,
                            "flip_to_wrong_pct": round(_flip(sub, c) * 100, 2), "n_orig_correct": len(sub)})
    return pd.DataFrame(over), pd.DataFrame(byt)


DRAFT_CONDS = ["all_wrong_cot", "mask_draft", "corrupt_draft", "swap_cot", "drop_cot"]


def draft_frames():
    """R2b draft-SMILES perturbation: flip-to-wrong on the subset originally-correct WITH a
    drafted SMILES (base_correct==1 & n_draft>0). Only the 3 models where it was run."""
    rows = []
    for stem, f in models_present("drift"):
        if stem not in ("Chem-R", "Chem-R-Faithful", "ChemDFM-R"):
            continue
        lab = LABEL[stem]; pe = json.load(open(f))["per_example"]
        bc = [e for e in pe if e.get("base_correct") == 1]
        draft = [e for e in bc if e.get("n_draft", 0) > 0]
        cov = round(len(draft) / len(bc) * 100, 1) if bc else None
        grp = lambda t: "translate" if t in TRANSLATE else "s2"
        for g in ["all", "translate", "s2"]:
            sub = draft if g == "all" else [e for e in draft if grp(e["task"]) == g]
            if not sub:
                continue
            row = {"model": lab, "task_group": g, "n_correct_with_draft": len(sub),
                   "coverage_pct_of_correct": cov if g == "all" else ""}
            for c in DRAFT_CONDS:
                row[c.replace("_cot", "").replace("_draft", "_draft") + "_flip_pct"] = round(_flip(sub, c) * 100, 2)
            rows.append(row)
    return pd.DataFrame(rows)


def cond_frames():
    cols = ["H_noCoT", "H_realCoT", "H_corrCoT", "ig_presence", "ig_content", "ig_swap"]
    summ, byt = [], []
    for stem, f in models_present("condsent"):
        c = json.load(open(f)); lab = LABEL[stem]; pt = c.get("per_task", {})
        def mot(k):
            vs = [v[k] for v in pt.values() if v.get(k) is not None]
            return round(sum(vs) / len(vs), 4) if vs else None
        summ.append({"model": lab, "aggregation": "mean_over_tasks (PLOTTED in R3)", "n_tasks": len(pt),
                     **{k: mot(k) for k in cols}})
        er = c.get("er_split", {})
        for k, lab2 in [("all", "pooled_all_examples"), ("ER=0", "pooled_ER=0"), ("ER>0", "pooled_ER>0")]:
            v = er.get(k)
            if v:
                summ.append({"model": lab, "aggregation": lab2, "n": v.get("n"),
                             **{x: (round(v[x], 4) if v.get(x) is not None else None) for x in cols}})
        for task, v in pt.items():
            byt.append({"model": lab, "task": task, "n": v.get("n"),
                        **{x: (round(v[x], 4) if v.get(x) is not None else None) for x in cols}})
    return pd.DataFrame(summ), pd.DataFrame(byt)


EXCLUDE_MODELS = set()  # no non-paper models shipped


def csv_df(name):
    p = os.path.join(D, name)
    df = pd.read_csv(p) if os.path.exists(p) else pd.DataFrame({"missing": [name]})
    if "model" in df.columns:
        df = df[~df["model"].isin(EXCLUDE_MODELS)]
    return df


d_over, d_byt = drift_frames()
d_draft = draft_frames()
c_summ, c_byt = cond_frames()

# README / sheet index
readme = pd.DataFrame([
    ["README", "this index + metric definitions + N", "", "—"],
    ["Diagnosis_model_task", "per (model,task): perf, ER=0/ER>0, 2x2, GC/CP, entropy", "test sets: cap2mol/mol2cap 3300, retro 5007, s2 500x9", "eval/export_stats.py"],
    ["Diagnosis_family", "per (model,family) sample-weighted aggregates", "same", "eval/export_stats.py"],
    ["R1_stage_ladder", "origin-of-hallucination ladder: per-claim fab, hedge/abstain, claims/resp", "n_resp 11,607 (3 GEN); answer-only=cap2mol 3300", "eval/stage_ladder_metrics.py"],
    ["R2_drift", "flip-to-wrong among ORIGINALLY-CORRECT, per condition x task-group (PAPER metric)", "n_orig_correct col (translate ~4-5k, s2 ~1.1-1.5k)", "eval/cot_drift.py -> drift_<m>.json (per_example)"],
    ["R2_drift_by_task", "same, per task", "n_orig_correct col", "eval/cot_drift.py"],
    ["R2_draft_perturbation", "DIRECT test of the structural-draft channel: mask/corrupt the drafted SMILES vs FG-name(all_wrong)/whole-CoT(swap); subset = correct WITH a draft; ChemDFM=neg ctrl", "n_correct_with_draft col", "eval/cot_drift.py (mask/corrupt_draft) -> eval/pull_draft.py"],
    ["R3_condentropy", "metric-free info-gains; mean_over_tasks row = PLOTTED, + pooled + ER=0/ER>0", "~13,560 x 8 samples", "eval/cot_condsent.py -> condsent_<m>.json"],
    ["R3_condentropy_task", "same, per task (granular; reproduces mean_over_tasks)", "n col", "eval/cot_condsent.py"],
    ["R4_mitigation", "baseline vs +process vs +coupled, per family", "per family", "eval/export_stats.py"],
    ["R5_grad_enrichment", "gradient x input saliency per token TYPE (share + enrichment)", "FULL ~16,100/model", "eval/attr_probe.py + eval/pull_fullvol.py"],
    ["R5_region_attention", "answer->token attention, INPUT vs TRACE region + draft-copy", "FULL ~16,100/model", "eval/attention_attribution.py --regions_only"],
    ["R5_token_examples", "per-token saliency+attention for the heatmap example figure", "12 curated examples (3 models x 4 tasks)", "eval/token_examples.py"],
], columns=["sheet", "contents", "N (data volume)", "regenerate"])

metric_notes = pd.DataFrame([
    ["ER / fabrication", "claimed FG absent from BOTH input and output by exact RDKit SMARTS; per-claim fab = Σfab/Σclaimed (generic FGs excluded)"],
    ["perf", "cap2mol/retro exact SMILES match; mol2cap caption similarity; s2 official S2-TOMG success"],
    ["flip_to_wrong (R2)", "among originally-correct (base_correct=1) responses, fraction that become incorrect (dperf=-1) after perturbing the reasoning; PAPER metric. NOTE: NOT the same as JSON summary.drift_rate (=any answer change over ALL examples). 95% bootstrap CIs are in RESULTS.md R2. wrong_input (input-FG control) is in RESULTS.md R2, computed separately (applies only where the FG is named in the input) — not in per_example, so omitted here."],
    ["ig_presence / ig_content / ig_swap (R3)", "H(noCoT)-H(realCoT) / H(corruptCoT)-H(realCoT) / H(swapCoT)-H(realCoT); H=Shannon over canonical-SMILES clusters of 8 samples; metric-free. R3 figure plots the mean_over_tasks aggregation (not the pooled-all-examples value)."],
    ["enrichment (R5)", "saliency_share / token_fraction; >1 = above-average answer-sensitivity per token of that type (fair metric; absolute share is count-confounded)"],
    ["draft_copy (R5)", "fraction of examples where the final answer SMILES appears (canonically) in the reasoning trace"],
    ["R2_draft_perturbation", "same flip-to-wrong metric, but perturbing ONLY the drafted SMILES: mask_draft = SMILES->[...] (presence); corrupt_draft = SMILES->valid-but-wrong structure (content). Subset = originally-correct examples whose trace drafts a SMILES (n_draft>0). Direct causal test that the structural draft, not the FG-name prose, is load-bearing; ChemDFM = negative control (only 8% of its correct traces draft any SMILES)."],
    ["CAVEAT", "R5 grad/attn are DESCRIPTIVE probes; causal claim carried by R2/R3. ChemDFM positive channel under-specified. base-a perf~0 -> drift uninformative. DeepSeek excluded (n=83)."],
], columns=["metric", "definition"])

with pd.ExcelWriter(XLSX, engine="openpyxl") as w:
    readme.to_excel(w, "README", index=False, startrow=1)
    metric_notes.to_excel(w, "README", index=False, startrow=len(readme) + 5)
    csv_df("stats_per_model_task.csv").to_excel(w, "Diagnosis_model_task", index=False)
    csv_df("stats_per_family.csv").to_excel(w, "Diagnosis_family", index=False)
    csv_df("stage_ladder.csv").to_excel(w, "R1_stage_ladder", index=False)
    d_over.to_excel(w, "R2_drift", index=False)
    d_byt.to_excel(w, "R2_drift_by_task", index=False)
    d_draft.to_excel(w, "R2_draft_perturbation", index=False)
    c_summ.to_excel(w, "R3_condentropy", index=False)
    c_byt.to_excel(w, "R3_condentropy_task", index=False)
    csv_df("mitigation.csv").to_excel(w, "R4_mitigation", index=False)
    csv_df("token_examples/r5_gradient.csv").to_excel(w, "R5_grad_enrichment", index=False)
    csv_df("token_examples/r5_region.csv").to_excel(w, "R5_region_attention", index=False)
    csv_df("token_examples/token_examples.csv").to_excel(w, "R5_token_examples", index=False)

# light formatting: bold header, freeze top row, autosize (cap 60)
from openpyxl import load_workbook
from openpyxl.styles import Font
wb = load_workbook(XLSX)
for ws in wb.worksheets:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
    # SMILES tokens like "=O"/"=C" start with '=' and openpyxl mis-stored them as FORMULAS
    # (invalid -> Excel repairs by deleting them). Force every such cell back to literal text.
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                cell.data_type = "s"
    for col in ws.columns:
        try:
            width = min(60, max(10, max(len(str(c.value)) for c in col if c.value is not None) + 2))
            ws.column_dimensions[col[0].column_letter].width = width
        except Exception:
            pass
wb.save(XLSX)
print(f"wrote {XLSX}")
print(f"sheets: {len(wb.sheetnames)} -> {wb.sheetnames}")
